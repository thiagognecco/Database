"""Import/Export endpoints for CSV and XLSX."""
from io import StringIO, BytesIO
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import csv

from app.database import get_db
from app.importador import importar_csv
from app.models import Link

router = APIRouter(prefix="/api", tags=["import-export"])


@router.post("/import")
async def import_csv_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import links from CSV file."""
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Apenas arquivos CSV são suportados")

    # Read file
    try:
        content = await file.read()
        csv_path = f"/tmp/{file.filename}"
        with open(csv_path, "wb") as f:
            f.write(content)

        result = importar_csv(csv_path, db=db)
        return result

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar CSV: {str(e)}")


@router.get("/export")
def export_links(format: str = "xlsx", db: Session = Depends(get_db)):
    """Export all links to CSV or XLSX."""
    if format not in ["csv", "xlsx"]:
        raise HTTPException(status_code=400, detail="Formato deve ser 'csv' ou 'xlsx'")

    links = db.query(Link).all()

    if format == "csv":
        return _export_csv(links)
    else:
        return _export_xlsx(links)


def _export_csv(links: list):
    """Export links to CSV."""
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "LINK",
            "Autor",
            "Data",
            "Plataforma",
            "Categoria",
            "Tema",
            "Título",
            "Resumo (20 palavras)",
            "Confiabilidade",
            "Bot",
        ],
    )

    writer.writeheader()

    for link in links:
        writer.writerow({
            "LINK": link.url,
            "Autor": link.autor or "",
            "Data": link.data.strftime("%d/%m/%Y") if link.data else "",
            "Plataforma": link.plataforma or "",
            "Categoria": link.categoria or "",
            "Tema": link.tema or "",
            "Título": link.titulo or "",
            "Resumo (20 palavras)": link.resumo or "",
            "Confiabilidade": link.confiabilidade or "Média",
            "Bot": "Sim" if link.bot else "Não",
        })

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=banco_de_links.csv"},
    )


def _export_xlsx(links: list):
    """Export links to XLSX."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl não instalado. Instale com: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Links"

    # Headers
    headers = [
        "LINK",
        "Autor",
        "Data",
        "Plataforma",
        "Categoria",
        "Tema",
        "Título",
        "Resumo (20 palavras)",
        "Confiabilidade",
        "Bot",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, link in enumerate(links, 2):
        ws.cell(row=row_num, column=1).value = link.url
        ws.cell(row=row_num, column=2).value = link.autor or ""
        ws.cell(row=row_num, column=3).value = (
            link.data.strftime("%d/%m/%Y") if link.data else ""
        )
        ws.cell(row=row_num, column=4).value = link.plataforma or ""
        ws.cell(row=row_num, column=5).value = link.categoria or ""
        ws.cell(row=row_num, column=6).value = link.tema or ""
        ws.cell(row=row_num, column=7).value = link.titulo or ""
        ws.cell(row=row_num, column=8).value = link.resumo or ""
        ws.cell(row=row_num, column=9).value = link.confiabilidade or "Média"
        ws.cell(row=row_num, column=10).value = "Sim" if link.bot else "Não"

    # Adjust column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 10

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=banco_de_links.xlsx"},
    )


@router.post("/backup")
def create_backup_endpoint(db: Session = Depends(get_db)):
    """Create a timestamped backup of database."""
    from app.database import create_backup
    result = create_backup()
    if result.get("success"):
        return {"message": "Backup criado com sucesso", "backup": result["backup"]}
    else:
        raise HTTPException(status_code=500, detail=result.get("error"))


@router.get("/backups")
def list_backups_endpoint(db: Session = Depends(get_db)):
    """List all available backups."""
    from app.database import list_backups
    return list_backups()


@router.post("/restore/{backup_name}")
def restore_backup_endpoint(backup_name: str, db: Session = Depends(get_db)):
    """Restore database from backup."""
    from app.database import restore_backup
    from pathlib import Path

    backup_path = Path(__file__).parent.parent.parent / "dados" / "backups" / backup_name
    result = restore_backup(str(backup_path))
    if result.get("success"):
        return result
    else:
        raise HTTPException(status_code=500, detail=result.get("error"))
